# coding=utf-8
import json
import time
import datetime

from flask import jsonify, request, render_template
from flask_pymongo import DESCENDING
from openpyxl import Workbook, load_workbook
import flask_excel as excel

from main import app, mongo
from main.router.lazy import check_login

vars_json = {'1': '现金',
             '2': '信用卡支付',
             '3': '银联卡支付',
             '4': '微信',
             '5': '支付宝',
             '6': '车主邦支付',
             '7': '滴滴支付',
             '8': '百度钱包',
             '9': '翼支付',
             '10': '车e族',
             '11': 'ETC支付',
             '12': '易捷支付',
             '13': '中经汇通',
             '14': '和信通',
             '15': '国通卡支付',
             '16': '银闪付',
             '17': '加油卡支付',
             '18': '会员卡支付'}
var_json_server = {
    '1': '便利店',
    '2': '卫生间',
    '3': '餐饮',
    '4': '住宿',
    '5': '银行ATM机',
    '6': '饮用热水',
    '7': '换油服务',
    '8': '换胎服务',
    '9': '汽车充气',
    '10': '维修保养',
    '11': '加油卡充值业务',
    '12': '银联卡充值',
    '13': '发卡充值网点',
    '14': '公交卡充值'
}


@app.route("/gas", methods=['GET'])
@check_login
def ex_gas():
    filters = {}
    tday = request.args.get('tday')
    if tday is None:
        tday = str(datetime.date.today())
    start_time = int(time.mktime(time.strptime(str(tday), '%Y-%m-%d')))
    ddd = datetime.datetime.strptime(tday, '%Y-%m-%d')
    tomorrow = ddd + datetime.timedelta(days=1)
    end_time = int(time.mktime(time.strptime(str(tomorrow.strftime('%Y-%m-%d')), '%Y-%m-%d')))
    filters['data'] = {'$gte': start_time, '$lte': end_time}
    gas = mongo.db.dingtalkgas.find(filters).sort('data', DESCENDING)
    gas_list = []
    count = mongo.db.dingtalkgas.find(filters).count()
    for g in gas:
        g.pop('_id')
        g['gas'] = json.loads(g['gas'])
        g['data'] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(int(g['data'])))
        gas_list.append(g)
    gas_list.sort(key=takeSecond)
    return render_template('gas/gas.html', list=gas_list, count=count, ttime=tday)


def takeSecond(elem):
    return elem['gas']['user']['name']


@app.route("/gas/excel", methods=['GET'])
@check_login
def ex_gas_to_excel():
    filters = {}
    tday = request.args.get('tday')
    if tday is None:
        tday = str(datetime.date.today())
    start_time = int(time.mktime(time.strptime(str(tday), '%Y-%m-%d')))
    ddd = datetime.datetime.strptime(tday, '%Y-%m-%d')
    tomorrow = ddd + datetime.timedelta(days=1)
    end_time = int(time.mktime(time.strptime(str(tomorrow.strftime('%Y-%m-%d')), '%Y-%m-%d')))
    filters['data'] = {'$gte': start_time, '$lte': end_time}
    gas = mongo.db.dingtalkgas.find(filters).sort('data', DESCENDING)
    gas_list = []
    var = ('录入时间',
           '录入人',
           '油站名称',
           '加油站类型',
           '加油站地址',
           '加油站经度',
           '加油站维度',
           '加油站联系人',
           '联系人电话',
           '座机电话 ',
           '-40#',
           '-35#',
           '-30#',
           '-20#',
           '-10#',
           '国四0 #',
           '0',
           'CNG',
           'LNG',
           '90#',
           '92#',
           '93#',
           '95#',
           '97#',
           '98#',
           '101#',
           '支付方式',
           '服务',
           '活动名称',
           '活动时间',
           '活动详情',
           '会员卡类别',
           '卡售卖时间',
           '卡优惠政策',
           '图片')
    gas_list.append(var)
    for g in gas:
        gas_msg = json.loads(g['gas'])
        user_msg = gas_msg['user']
        petrolStation = gas_msg['petrolStation']
        act = gas_msg['activity']
        g_ = (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(int(g['data']))), user_msg['name'],
              petrolStation['petrolStationName'], petrolStation['petrolStationType'],
              petrolStation['petrolStationSite'], str(petrolStation['lng']), str(petrolStation['lat']),
              petrolStation['petrolStationUserName'], petrolStation['petrolStationPhone'],
              petrolStation['petrolStationSpecialPhone'], oil(petrolStation, '1'), oil(petrolStation, '2'),
              oil(petrolStation, '3'), oil(petrolStation, '4'), oil(petrolStation, '5'), oil(petrolStation, '6'),
              oil(petrolStation, '7'), oil(petrolStation, '8'), oil(petrolStation, '9'), oil(petrolStation, '10'),
              oil(petrolStation, '11'), oil(petrolStation, '12'), oil(petrolStation, '13'), oil(petrolStation, '14'),
              oil(petrolStation, '15'), oil(petrolStation, '16'), get_value(gas_msg), get_value_server(gas_msg),
              act['livename'], act['livetime'], act['liveinfo'], act['vipcat'], act['vipcattime'],
              act['vipcatdiscounts'], for2Str(gas_msg['imgList']))
        gas_list.append(g_)
    return to_excel(gas_list)


def for2Str(josn):
    string_ = ''
    for js in josn:
        string_ + " " + str(js)
    return string_


def get_value_server(json_str):
    pay = json_str['Serve']
    pay_str = ''
    for p in pay:
        pay_str + " " + var_json_server[p]
    return pay_str


def get_value(json_str):
    pay = json_str['pay']
    pay_str = ''
    for p in pay:
        pay_str + " " + vars_json[p]
    return pay_str


def oil(josn_str, num):
    oilList = josn_str['oilList']
    for oil in oilList:
        if oil['Ykey'] == num:
            return oil['val']
    else:
        return ''


def to_excel(query_sets):
    return excel.make_response_from_array(
        query_sets,
        file_type='xlsx',
        file_name='list.xlsx'
    )


def json2xls(data):
    wb1 = Workbook()
    sheet = wb1.active
    listHead = []
    for c, i in enumerate(data[0].keys()):
        sheet.cell(row=1, column=c + 1, value=i)
        listHead.append(i)
    for r, i in enumerate(data):
        print r, i
        row = r + 2
        for c, d in enumerate(listHead):
            print c, d, i.get(d, "")
            if d == 'gas':
                continue
            sheet.cell(row=row, column=c + 1, value=i.get(d, ""))
    wb1.save('test' + ".xlsx")
